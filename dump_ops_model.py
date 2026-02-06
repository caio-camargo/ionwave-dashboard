"""
Dump all sheets from IonWave_Ops_Model_v10_Protocol.xlsx into JSON files.
One JSON file per sheet, plus a _manifest.json summarizing all sheets.
"""
import openpyxl
import json
import os
import re

XLSX_PATH = "IonWave_Ops_Model_v10_Protocol.xlsx"
OUTPUT_DIR = "ops_model_v10_dump"

def slugify(name):
    """Convert sheet name to a filesystem-safe slug."""
    s = name.strip().lower()
    s = re.sub(r'[^a-z0-9]+', '_', s)
    s = s.strip('_')
    return s

def cell_value(cell):
    """Extract cell value, handling None and types."""
    v = cell.value
    if v is None:
        return None
    if isinstance(v, (int, float, bool)):
        return v
    return str(v).strip()

def dump_sheet(ws):
    """Dump a worksheet to a list of row-dicts using first row as headers."""
    rows = list(ws.iter_rows())
    if not rows:
        return {"headers": [], "rows": [], "row_count": 0}

    # Find the first non-empty row to use as headers
    header_row_idx = 0
    for i, row in enumerate(rows):
        vals = [cell_value(c) for c in row]
        if any(v is not None and v != '' for v in vals):
            header_row_idx = i
            break

    raw_headers = [cell_value(c) for c in rows[header_row_idx]]
    # Create headers, using column index for empty ones
    headers = []
    for j, h in enumerate(raw_headers):
        if h is None or str(h).strip() == '':
            headers.append(f"col_{j+1}")
        else:
            headers.append(str(h))

    # Extract data rows
    data_rows = []
    for row in rows[header_row_idx + 1:]:
        vals = [cell_value(c) for c in row]
        # Skip completely empty rows
        if not any(v is not None and v != '' for v in vals):
            continue
        row_dict = {}
        for j, val in enumerate(vals):
            if j < len(headers):
                row_dict[headers[j]] = val
            else:
                row_dict[f"col_{j+1}"] = val
        data_rows.append(row_dict)

    return {
        "headers": headers,
        "rows": data_rows,
        "row_count": len(data_rows)
    }

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print(f"Loading {XLSX_PATH}...")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)

    manifest = {
        "source_file": XLSX_PATH,
        "sheet_count": len(wb.sheetnames),
        "sheets": []
    }

    for idx, name in enumerate(wb.sheetnames):
        slug = slugify(name)
        filename = f"{idx+1:02d}_{slug}.json"
        filepath = os.path.join(OUTPUT_DIR, filename)

        print(f"  [{idx+1}/{len(wb.sheetnames)}] {name} -> {filename}")

        ws = wb[name]
        sheet_data = dump_sheet(ws)
        sheet_data["sheet_name"] = name
        sheet_data["sheet_index"] = idx + 1

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(sheet_data, f, indent=2, ensure_ascii=False, default=str)

        manifest["sheets"].append({
            "index": idx + 1,
            "name": name,
            "slug": slug,
            "filename": filename,
            "row_count": sheet_data["row_count"],
            "header_count": len(sheet_data["headers"]),
            "headers": sheet_data["headers"]
        })

        print(f"    -> {sheet_data['row_count']} rows, {len(sheet_data['headers'])} columns")

    # Write manifest
    manifest_path = os.path.join(OUTPUT_DIR, "_manifest.json")
    with open(manifest_path, 'w', encoding='utf-8') as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)

    print(f"\nDone! {len(wb.sheetnames)} sheets dumped to {OUTPUT_DIR}/")
    print(f"Manifest: {manifest_path}")

    # Print summary
    total_rows = sum(s["row_count"] for s in manifest["sheets"])
    print(f"Total rows across all sheets: {total_rows}")

    wb.close()

if __name__ == "__main__":
    main()
